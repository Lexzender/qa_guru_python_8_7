from zipfile import ZipFile
import pandas
from openpyxl import load_workbook
from pypdf import PdfReader
from conftest import file_archiving

my_file = ["file_example_XLSX_50.xlsx", "file_example_XLS_10.xls", "hello.zip",
           "Python Testing with Pytest (Brian Okken) (1).pdf", "text.txt"]


def test_file(dell_zip):
    # Создаем архив с файлами
    with ZipFile('tmp/Python.zip', 'w') as zip_File:
        file_archiving('resources/', zip_File)

    # Проверяем, что созданный архив содержит нужные документы
    with ZipFile("tmp/Python.zip") as zip_File:
        info = zip_File.infolist()
        for i in info:
            assert i.filename in my_file, f"{i.filename}: Файл не найден в архиве"


def test_count_file(dell_zip):
    # Создаем архив с файлами
    with ZipFile('tmp/Python.zip', 'w') as zip_File:
        file_archiving('resources/', zip_File)

    # Проверяем количество файлов в архиве
    with ZipFile("tmp/Python.zip") as zip_File:
        info = zip_File.namelist()
        assert len(info) == 5, f"Количество файлов в архиве != 5"


def test_pdf_file(dell_zip):
    # Создаем архив с файлами
    with ZipFile('tmp/Python.zip', 'w') as zip_File:
        file_archiving('resources/', zip_File)

    # Проверяем содержимое pdf
    with ZipFile("tmp/Python.zip") as zip_File:
        with zip_File.open("Python Testing with Pytest (Brian Okken) (1).pdf") as pdf_f:
            pdf_file = PdfReader(pdf_f)
            assert "Python Testing with pytest" in pdf_file.pages[1].extract_text()
            number_of_page = len(pdf_file.pages)
            assert number_of_page == 256


def test_xlsx_file(dell_zip):
    # Создаем архив с файлами
    with ZipFile('tmp/Python.zip', 'w') as zip_File:
        file_archiving('resources/', zip_File)

    # Проверяем содержимое xlsx
    with ZipFile("tmp/Python.zip") as zip_File:
        with zip_File.open("file_example_XLSX_50.xlsx") as xlsx_f:
            xlsx_file = load_workbook(xlsx_f)
            sheet = xlsx_file.active
            sheet_value = sheet.cell(row=2, column=2).value
            assert sheet_value == 'Dulce'
            num_rows = sheet.max_row
            num_cols = sheet.max_column
            assert num_rows == 51
            assert num_cols == 8


def test_xls_file(dell_zip):
    # Создаем архив с файлами
    with ZipFile('tmp/Python.zip', 'w') as zip_File:
        file_archiving('resources/', zip_File)

    # Проверяем содержимое xls
    with ZipFile("tmp/Python.zip") as zip_File:
        with zip_File.open("file_example_XLS_10.xls") as xls_f:
            table = pandas.read_excel(xls_f).head(2)
            assert 'Hashimoto' in table['Last Name'].values


def test_txt_file(dell_zip):
    # Создаем архив с файлами
    with ZipFile('tmp/Python.zip', 'w') as zip_File:
        file_archiving('resources/', zip_File)

    # Проверяем содержимое txt
    with ZipFile("tmp/Python.zip") as zip_File:
        with zip_File.open("text.txt") as txt_f:
            assert txt_f.read().decode('utf-8') == 'текст для теста'
